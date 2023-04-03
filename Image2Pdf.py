import openpyxl
from PIL import Image, ImageFont, ImageDraw

myFont = ImageFont.truetype("Tangerine-Bold.ttf", 35)
file = "Book1.xlsx"
work = openpyxl.load_workbook(file)
sheet = work.active
names = []
print("Reading file...")

for i in range(0, sheet.max_row):
    for col in sheet.iter_cols(1, sheet.max_column):
        names.append(col[i].value)


print("Reading Completed.")

for name in names:
    img = Image.open("Final.jpg")
    draw = ImageDraw.Draw(img)
    w, h = draw.textsize(name, font=myFont)
    draw.text((244+(512-w)/2, 355), name, fill="black", font=myFont)
    img.save("Final\\" + name + ".pdf")
