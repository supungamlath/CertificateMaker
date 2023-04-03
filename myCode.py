import openpyxl
from PyPDF2 import PdfFileWriter, PdfFileReader
import io

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


pdfmetrics.registerFont(TTFont('Tangerine', 'Tangerine-Bold.ttf'))
file = "Book1.xlsx"
work = openpyxl.load_workbook(file)
sheet = work.active
names= []
print("Reading file...")

for i in range(0, sheet.max_row):
    for col in sheet.iter_cols(1, sheet.max_column):
        names.append(col[i].value)



print("Reading Completed.")

print('Creating pdf files...')


for name in names :
    addition = io.BytesIO()
    newcan = canvas.Canvas(addition)
    newcan.setFont('Tangerine', 35)
    newcan.drawCentredString(430, 285, name,6)
    newcan.save()

    addition.seek(0)

    myNew = PdfFileReader(addition)

    existing = PdfFileReader(open("Final.pdf", "rb"))

    output = PdfFileWriter()

    page = existing.getPage(0)
    page.mergePage(myNew.getPage(0))
    output.addPage(page)

    stream = open("Final/"+name + ".pdf", "wb")
    output.write(stream)
    stream.close()


print('Task Completed.')
